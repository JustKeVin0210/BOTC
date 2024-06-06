import os

import json
import math
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import PatternFill, Alignment, Side, Border, Font


def write_excel(json_file, image_dir, save_dir):
    os.makedirs(save_dir, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    side_style = None
    format_dict = {
        "title": {
            "fill": PatternFill("solid", fgColor="FF000000"),
            "alignment": Alignment(horizontal='center', vertical='center'),
            "border": Border(left=Side(style=side_style),
                             bottom=Side(style=side_style),
                             right=Side(style=side_style),
                             top=Side(style=side_style)),
            "font": Font(name='宋体', bold=True, color="FFFFFFFF"),
        },
        "time": {
            "fill": PatternFill("solid", fgColor="FF00FF00"),
            "alignment": Alignment(horizontal='center', vertical='center'),
            "border": Border(left=Side(style=side_style),
                             bottom=Side(style=side_style),
                             right=Side(style=side_style),
                             top=Side(style=side_style)),
            "font": Font(name='宋体', italic=True, color="FF000000"),
        },
        "townsfolk": {
            "team_cn": "镇民",
            "fill": PatternFill("solid", fgColor="FF0070C0"),
            "alignment": Alignment(horizontal='center', vertical='center'),
            "border": Border(left=Side(style=side_style),
                             bottom=Side(style=side_style),
                             right=Side(style=side_style),
                             top=Side(style=side_style)),
            "font": Font(name='宋体', color="FF000000"),
        },
        "outsider": {
            "team_cn": "外来者",
            "fill": PatternFill("solid", fgColor="FF00B0F0"),
            "alignment": Alignment(horizontal='center', vertical='center'),
            "border": Border(left=Side(style=side_style),
                             bottom=Side(style=side_style),
                             right=Side(style=side_style),
                             top=Side(style=side_style)),
            "font": Font(name='宋体', color="FF000000"),
        },
        "minion": {
            "team_cn": "爪牙",
            "fill": PatternFill("solid", fgColor="FFFFC000"),
            "alignment": Alignment(horizontal='center', vertical='center'),
            "border": Border(left=Side(style=side_style),
                             bottom=Side(style=side_style),
                             right=Side(style=side_style),
                             top=Side(style=side_style)),
            "font": Font(name='宋体', color="FF000000"),
        },
        "demon": {
            "team_cn": "恶魔",
            "fill": PatternFill("solid", fgColor="FFFF0000"),
            "alignment": Alignment(horizontal='center', vertical='center'),
            "border": Border(left=Side(style=side_style),
                             bottom=Side(style=side_style),
                             right=Side(style=side_style),
                             top=Side(style=side_style)),
            "font": Font(name='宋体', color="FF000000"),
        },
        "traveler": {
            "team_cn": "旅行者",
            "fill": PatternFill("solid", fgColor="FF7030A0"),
            "alignment": Alignment(horizontal='center', vertical='center'),
            "border": Border(left=Side(style=side_style),
                             bottom=Side(style=side_style),
                             right=Side(style=side_style),
                             top=Side(style=side_style)),
            "font": Font(name='宋体', color="FFFFFFFF"),
        },
        "fabled": {
            "team_cn": "传奇角色",
            "fill": PatternFill("solid", fgColor="FFFFFF00"),
            "alignment": Alignment(horizontal='center', vertical='center'),
            "border": Border(left=Side(style=side_style),
                             bottom=Side(style=side_style),
                             right=Side(style=side_style),
                             top=Side(style=side_style)),
            "font": Font(name='宋体', color="FF000000"),
        },
    }

    ws.title = "全角色"

    first_line = [None, "角色名", "角色类型", "剧本", "首夜行动", None, "角色名", "角色类型", "剧本", "其他夜行动"]
    ws.append(first_line)
    data_len = len(first_line) // 2

    with open(json_file, "r", encoding='utf-8') as f:
        script_roles = json.load(f)
    if script_roles[0].get("id") == "_meta":
        script_roles.pop(0)
    script_times = [
        {
            "id": "duskTRANS",
            "name": "黄昏",
            "team": "time",
            "firstNight": 1,
            "otherNight": 1,
        },
        {
            "id": "minionTRANS",
            "name": "爪牙信息",
            "team": "minion",
            "firstNight": 2000,
            "otherNight": 0,
        },
        {
            "id": "demonTRANS",
            "name": "恶魔信息",
            "team": "demon",
            "firstNight": 3000,
            "otherNight": 0,
        },
        {
            "id": "dawnTRANS",
            "name": "黎明",
            "team": "time",
            "firstNight": 12600,
            "otherNight": 15000,
        }
    ]
    script_roles += script_times
    script_roles_fisrt_night = sorted([i for i in script_roles if i.get("firstNight")],
                                      key=lambda x: x.get("firstNight"))
    script_roles_other_night = sorted([i for i in script_roles if i.get("otherNight")],
                                      key=lambda x: x.get("otherNight"))
    if len(script_roles_fisrt_night) < len(script_roles_other_night):
        script_roles_fisrt_night += [None] * (len(script_roles_other_night) - len(script_roles_fisrt_night))
    else:
        script_roles_other_night += [None] * (len(script_roles_fisrt_night) - len(script_roles_other_night))

    for i in range(len(first_line)):
        if i == 0 or i == data_len:
            ws.column_dimensions[chr(65 + i)].width = 2.5
        else:
            ws.column_dimensions[chr(65 + i)].width = 13
    for i in range(len(script_roles_fisrt_night) + 1):
        ws.row_dimensions[i + 1].height = 15

    role_first_team_list, role_other_team_list, merge_list = [], [], []
    for i, (role_first, role_other) in enumerate(zip(script_roles_fisrt_night, script_roles_other_night)):
        if role_first:
            if role_first.get("script"):
                role_first_infolist = [None, role_first.get("name"),
                                       format_dict.get(role_first.get("team")).get("team_cn"),
                                       role_first.get("script"), role_first.get("firstNight")]
                id_first = role_first.get("id")
                image_first = ExcelImage(os.path.join(image_dir, id_first.replace("TRANS", "").capitalize() + ".png"))
            else:
                role_first_infolist = [None, role_first.get("name"), None, None,
                                       math.ceil(role_first.get("firstNight") / 100) * 100]
                merge_list.append((f"B{i + 2}", f"D{i + 2}"))
                id_first = role_first.get("id")
                image_first = ExcelImage(os.path.join(image_dir, id_first.replace("TRANS", "").upper() + ".png"))
            image_first.width, image_first.height = 19, 19
            ws.add_image(image_first, f"{chr(65)}{i + 2}")
            role_first_team_list.append(role_first.get("team"))
        else:
            role_first_team_list.append(None)
            role_first_infolist = [None] * data_len
        if role_other:
            if role_other.get("script"):
                role_other_infolist = [None, role_other.get("name"),
                                       format_dict.get(role_other.get("team")).get("team_cn"),
                                       role_other.get("script"), role_other.get("otherNight")]
                id_other = role_other.get("id")
                image_other = ExcelImage(os.path.join(image_dir, id_other.replace("TRANS", "").capitalize() + ".png"))
            else:
                role_other_infolist = [None, role_other.get("name"), None, None,
                                       math.ceil(role_other.get("otherNight") / 100) * 100]
                merge_list.append((f"G{i + 2}", f"I{i + 2}"))
                id_other = role_other.get("id")
                image_other = ExcelImage(os.path.join(image_dir, id_other.replace("TRANS", "").upper() + ".png"))
            image_other.width, image_other.height = 19, 19
            ws.add_image(image_other, f"{chr(65 + data_len)}{i + 2}")
            role_other_team_list.append(role_other.get("team"))
        else:
            role_other_team_list.append(None)
            role_other_infolist = [None] * data_len

        ws.append(role_first_infolist + role_other_infolist)

    # 格式修改
    # 表头设置
    for cell in ws[1]:
        cell.fill = format_dict.get("title").get("fill")
        cell.alignment = format_dict.get("title").get("alignment")
        cell.border = format_dict.get("title").get("border")
        cell.font = format_dict.get("title").get("font")

    # 从第二行开始设置
    for row, role_first_team, role_other_team in zip(ws.iter_rows(min_row=2),
                                                     role_first_team_list, role_other_team_list):
        for i, cell in enumerate(row):
            if role_first_team and i < data_len:
                cell.fill = format_dict.get(role_first_team).get("fill")
                cell.alignment = format_dict.get(role_first_team).get("alignment")
                cell.border = format_dict.get(role_first_team).get("border")
                cell.font = format_dict.get(role_first_team).get("font")
            if role_other_team and i >= data_len:
                cell.fill = format_dict.get(role_other_team).get("fill")
                cell.alignment = format_dict.get(role_other_team).get("alignment")
                cell.border = format_dict.get(role_other_team).get("border")
                cell.font = format_dict.get(role_other_team).get("font")

    # 合并单元格
    for merge_position in merge_list:
        ws.merge_cells(range_string=f"{merge_position[0]}:{merge_position[1]}")

    # 修改打印样式
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_margins.left = 0
    ws.page_margins.right = 0
    ws.page_margins.top = 0
    ws.page_margins.bottom = 0
    ws.page_margins.header = 0
    ws.page_margins.footer = 0
    ws.page_setup.fitToPage = True
    # 保存
    wb.save(os.path.join(save_dir, "角色行动顺序表.xlsx"))
    print(f"写入{os.path.join(save_dir, '角色行动顺序表.xlsx')}完成")
    return


if __name__ == '__main__':
    json_file = r"json/全角色.json"
    image_dir = r"image_all"
    save_dir = r"./"
    write_excel(json_file, image_dir, save_dir)
